from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from app.numbers import is_blank, normalize_uk_number

PHONE_HEADER_EXACT = {
    "phone",
    "telephone",
    "mobile",
    "landline",
    "tel",
    "cell",
    "cellphone",
    "msisdn",
    "ddi",
    "switchboard",
    "directdial",
    "directline",
    "workphone",
    "homephone",
    "mobilephone",
    "mobilenumber",
    "phonenumber",
    "telephonenumber",
    "contactnumber",
    "contacttel",
    "contactphone",
}

PHONE_HEADER_PARTS = (
    "phone",
    "telephone",
    "mobile",
    "landline",
    "cellphone",
    "msisdn",
    "directdial",
    "directline",
    "ddi",
)

NEGATIVE_HEADERS = {
    "id",
    "email",
    "e-mail",
    "postcode",
    "post_code",
    "zip",
    "date",
    "amount",
    "price",
    "name",
    "firstname",
    "first_name",
    "lastname",
    "last_name",
    "fullname",
    "address",
    "city",
    "town",
    "county",
    "country",
    "company",
    "website",
    "url",
    "notes",
    "comment",
    "comments",
    "owner",
    "created",
    "updated",
    "status",
    "title",
    "job",
    "jobtitle",
    "invoice",
    "account",
    "reference",
    "ref",
    "value",
    "balance",
    "description",
    "source",
}


@dataclass
class ParsedRow:
    source_row: int
    source_field: str
    original: str
    normalized: str | None
    fields: dict[str, str] = field(default_factory=dict)


@dataclass
class ParsedFile:
    headers: list[str]
    phone_fields: list[str]
    items: list[ParsedRow]
    source_rows: int


def re_header(value: str) -> str:
    return "".join(ch for ch in (value or "").strip().lower() if ch.isalnum() or ch == "_")


def header_phone_score(header: str) -> int:
    raw = (header or "").strip().lower()
    key = re_header(header)
    if not key:
        return 0
    if key in PHONE_HEADER_EXACT:
        return 100
    if any(part in key for part in PHONE_HEADER_PARTS):
        if any(bad in key for bad in ("email", "postcode", "invoice", "account")):
            return 0
        return 80
    if re.search(r"\b(tel|mob|ddi)\b", raw):
        return 70
    return 0


def header_is_negative(header: str) -> bool:
    key = re_header(header)
    raw = (header or "").strip().lower()
    if key in NEGATIVE_HEADERS:
        return True
    return any(
        token in raw.split()
        for token in ("email", "postcode", "address", "website", "company", "owner")
    )


def _cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _looks_like_phone(value: str) -> bool:
    return normalize_uk_number(value) is not None


def detect_phone_columns(headers: list[str], rows: list[list[str]]) -> list[int]:
    width = max(len(headers), max((len(row) for row in rows), default=0))
    sample = rows[:80]
    scored: list[tuple[int, int]] = []

    for idx in range(width):
        header = headers[idx] if idx < len(headers) else ""
        values = [
            row[idx]
            for row in sample
            if idx < len(row) and not is_blank(row[idx])
        ]
        hits = sum(1 for value in values if _looks_like_phone(value))
        ratio = (hits / len(values)) if values else 0.0
        name_score = header_phone_score(header)

        if header_is_negative(header) and name_score == 0:
            continue
        if name_score >= 70 and hits >= 1:
            scored.append((idx, name_score + hits * 2))
        elif name_score >= 40 and ratio >= 0.3 and hits >= 1:
            scored.append((idx, name_score + hits * 2))
        elif name_score == 0 and ratio >= 0.5 and hits >= 2:
            scored.append((idx, hits * 2))

    if scored:
        scored.sort(key=lambda item: (-item[1], item[0]))
        return [idx for idx, _ in scored]

    best_idx = -1
    best_hits = 0
    for idx in range(width):
        hits = sum(
            1
            for row in sample
            if idx < len(row) and _looks_like_phone(row[idx])
        )
        if hits > best_hits:
            best_hits = hits
            best_idx = idx
    return [best_idx] if best_idx >= 0 and best_hits else []


def _looks_like_header_row(row: list[str]) -> bool:
    if not row:
        return False
    if any(header_phone_score(cell) or header_is_negative(cell) for cell in row):
        return True
    return not all(_looks_like_phone(cell) or is_blank(cell) for cell in row)


def _rows_from_csv(content: bytes) -> tuple[list[str] | None, list[list[str]]]:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = [row for row in reader if any(_cell(cell) for cell in row)]
    rows = [[_cell(cell) for cell in row] for row in rows]
    if not rows:
        return None, []
    if _looks_like_header_row(rows[0]):
        return [cell or f"column_{idx + 1}" for idx, cell in enumerate(rows[0])], rows[1:]
    return None, rows


def _rows_from_xlsx(content: bytes) -> tuple[list[str] | None, list[list[str]]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        values = [_cell(cell) for cell in row]
        if any(values):
            rows.append(values)
    workbook.close()
    if not rows:
        return None, []
    if _looks_like_header_row(rows[0]):
        return [cell or f"column_{idx + 1}" for idx, cell in enumerate(rows[0])], rows[1:]
    return None, rows


def _rows_from_txt(content: bytes) -> tuple[list[str] | None, list[list[str]]]:
    text = content.decode("utf-8-sig", errors="replace")
    rows: list[list[str]] = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if "," in line or "\t" in line:
            parts = [_cell(part) for part in re_split_line(line)]
            rows.append(parts)
        else:
            rows.append([line])
    if not rows:
        return None, []
    if len(rows[0]) > 1 and _looks_like_header_row(rows[0]):
        return [cell or f"column_{idx + 1}" for idx, cell in enumerate(rows[0])], rows[1:]
    return None, rows


def re_split_line(line: str) -> list[str]:
    if "\t" in line:
        return line.split("\t")
    return next(csv.reader([line]))


def _default_headers(width: int) -> list[str]:
    if width <= 1:
        return ["phone"]
    return [f"column_{idx + 1}" for idx in range(width)]


def parse_number_file(filename: str, content: bytes) -> ParsedFile:
    suffix = Path(filename).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        headers, rows = _rows_from_xlsx(content)
    elif suffix in {".csv"}:
        headers, rows = _rows_from_csv(content)
    else:
        headers, rows = _rows_from_txt(content)

    if not rows:
        return ParsedFile(headers=[], phone_fields=[], items=[], source_rows=0)

    width = max(len(row) for row in rows)
    if headers:
        while len(headers) < width:
            headers.append(f"column_{len(headers) + 1}")
    else:
        headers = _default_headers(width)

    phone_idxs = detect_phone_columns(headers, rows)
    phone_fields = [headers[idx] for idx in phone_idxs]
    items: list[ParsedRow] = []

    for offset, row in enumerate(rows, start=2 if headers else 1):
        fields = {
            header: (row[idx] if idx < len(row) else "")
            for idx, header in enumerate(headers)
            if header
        }
        found: list[ParsedRow] = []
        for idx in phone_idxs:
            header = headers[idx]
            raw = row[idx] if idx < len(row) else ""
            if is_blank(raw):
                continue
            found.append(
                ParsedRow(
                    source_row=offset,
                    source_field=header,
                    original=raw,
                    normalized=normalize_uk_number(raw),
                    fields=fields,
                )
            )
        if found:
            items.extend(found)
        else:
            items.append(
                ParsedRow(
                    source_row=offset,
                    source_field=phone_fields[0] if phone_fields else "",
                    original="",
                    normalized=None,
                    fields=fields,
                )
            )

    return ParsedFile(
        headers=headers,
        phone_fields=phone_fields,
        items=items,
        source_rows=len(rows),
    )
